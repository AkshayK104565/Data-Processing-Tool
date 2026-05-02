"""
Pattern Data Processing Automation Tool — Python Backend (Production-ready)
============================================================================
Deployment: Render / Railway / Azure App Service
Frontend:   MacroHub.html hosted on GitHub Pages / SharePoint

Routes:
  GET  /api/health                — health check
  POST /api/resolve-filenames     — batch URL → filename lookup
  POST /api/process-image-stack   — full xlsx upload → process → download
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import requests as req_lib
import re, io, os, urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ── CORS ────────────────────────────────────────────────────────────────────
# In production, restrict to your actual frontend origin.
# Set ALLOWED_ORIGIN env var on Render/Railway/Azure to your GitHub Pages URL
# e.g. https://yourname.github.io  or  https://yourcompany.sharepoint.com
# Leaving it as "*" works for testing but is less secure.
ALLOWED_ORIGIN = os.environ.get("ALLOWED_ORIGIN", "*")
CORS(app, origins=ALLOWED_ORIGIN)

# ── REQUEST TIMEOUT (seconds) ────────────────────────────────────────────────
URL_TIMEOUT = int(os.environ.get("URL_TIMEOUT", "15"))

# ── BROWSER HEADERS for outbound requests ───────────────────────────────────
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122 Safari/537.36"
    ),
    "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

# ────────────────────────────────────────────────────────────────────────────
#  FILENAME HELPERS
# ────────────────────────────────────────────────────────────────────────────
def extract_cd_filename(cd: str) -> str:
    """Extract filename from Content-Disposition header (RFC 5987 + plain)."""
    if not cd:
        return ""
    # RFC 5987  filename*=UTF-8''percent-encoded
    m = re.search(r"filename\*\s*=\s*(?:[^']*'')?([^;]+)", cd, re.IGNORECASE)
    if m:
        return urllib.parse.unquote(m.group(1).strip().strip("\"'"))
    # Plain     filename="name.jpg"
    m = re.search(r'filename\s*=\s*["\']?([^"\';\r\n]+)["\']?', cd, re.IGNORECASE)
    if m:
        return m.group(1).strip().strip("\"'")
    return ""

def sanitize(name: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]', "_", str(name))
    return name.rstrip(" .") or "download"

def fallback_from_url(url: str) -> str:
    """Extract filename from URL path — used when no Content-Disposition."""
    parsed = urllib.parse.urlparse(url)
    seg = parsed.path.rsplit("/", 1)[-1] if "/" in parsed.path else parsed.path
    return sanitize(urllib.parse.unquote(seg.split("?")[0])) or "download"

def resolve_filename(url: str) -> str:
    """
    HEAD request → Content-Disposition filename.
    Falls back to URL path if header absent.
    Exactly what a browser 'Save As' dialog does.
    """
    try:
        resp = req_lib.head(url, headers=HEADERS, allow_redirects=True, timeout=URL_TIMEOUT)
        fn = extract_cd_filename(resp.headers.get("Content-Disposition", ""))
        if fn:
            return sanitize(fn)
        return fallback_from_url(resp.url or url)
    except Exception:
        return fallback_from_url(url)


# ────────────────────────────────────────────────────────────────────────────
#  ROUTE 1  GET /api/health
# ────────────────────────────────────────────────────────────────────────────
@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "service": "Pattern Data Processing Backend"})


# ────────────────────────────────────────────────────────────────────────────
#  ROUTE 2  POST /api/resolve-filenames
#  Body: { "urls": ["https://...", ...] }   (max 200)
#  Returns: { "results": { "url": "filename.jpg", ... } }
# ────────────────────────────────────────────────────────────────────────────
@app.route("/api/resolve-filenames", methods=["POST"])
def api_resolve_filenames():
    data = request.get_json(force=True)
    urls = data.get("urls", [])
    if not urls:
        return jsonify({"error": "No URLs provided"}), 400
    if len(urls) > 200:
        return jsonify({"error": "Max 200 URLs per batch"}), 400

    results = {}
    with ThreadPoolExecutor(max_workers=10) as pool:
        fut_map = {pool.submit(resolve_filename, u): u for u in urls}
        for fut in as_completed(fut_map):
            url = fut_map[fut]
            try:
                results[url] = fut.result()
            except Exception:
                results[url] = fallback_from_url(url)
    return jsonify({"results": results})


# ────────────────────────────────────────────────────────────────────────────
#  ROUTE 3  POST /api/process-image-stack
#  Multipart: field "file" = .xlsx
#  Image Links sheet layout:
#    Col A = Master ID  |  Col B = Image Stack Group  |  Col C+ = image URLs
#  Returns: processed .xlsx download
# ────────────────────────────────────────────────────────────────────────────
@app.route("/api/process-image-stack", methods=["POST"])
def api_process_image_stack():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        wb_in = openpyxl.load_workbook(io.BytesIO(request.files["file"].read()), data_only=True)
    except Exception as e:
        return jsonify({"error": f"Could not read Excel file: {e}"}), 400

    if "Image Links" not in wb_in.sheetnames:
        return jsonify({"error": "Sheet 'Image Links' not found in uploaded file."}), 400

    ws_in = wb_in["Image Links"]
    rows  = list(ws_in.iter_rows(min_row=2, values_only=True))

    if not rows:
        return jsonify({"error": "No data rows in 'Image Links' (starting Row 2)."}), 400

    # Collect unique URLs + parsed row data
    all_urls, parsed_rows = set(), []
    for row in rows:
        mid = str(row[0] or "").strip() if len(row) > 0 else ""
        grp = str(row[1] or "").strip() if len(row) > 1 else ""
        if not mid or not grp:
            continue
        urls = [str(v or "").strip() for v in row[2:]]
        parsed_rows.append((mid, grp, urls))
        all_urls.update(u for u in urls if u)

    if not all_urls:
        return jsonify({"error": "No image URLs found in Columns C+ of 'Image Links'."}), 400

    # Resolve all filenames in parallel (12 threads)
    url_fn = {}
    with ThreadPoolExecutor(max_workers=12) as pool:
        fut_map = {pool.submit(resolve_filename, u): u for u in all_urls}
        for fut in as_completed(fut_map):
            url = fut_map[fut]
            url_fn[url] = fut.result() if not fut.exception() else fallback_from_url(url)

    # Build output workbook
    wb_out  = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    OUT_SN  = "Image Stack Import Template"

    # Copy original sheets
    for sn in wb_in.sheetnames:
        if sn == OUT_SN:
            continue
        ws_dst = wb_out.create_sheet(title=sn)
        for r in wb_in[sn].iter_rows(values_only=True):
            ws_dst.append(list(r))

    # Create output sheet
    ws_out = wb_out.create_sheet(title=OUT_SN)
    HDR = ["Import Type", "Collection Folder", "Image Stack Group", "Filename", "Image Stack Order"]
    hfont = Font(bold=True)
    hfill = PatternFill("solid", fgColor="BDD7EE")
    ws_out.append(HDR)
    for c, _ in enumerate(HDR, 1):
        cell = ws_out.cell(1, c)
        cell.font  = hfont
        cell.fill  = hfill
        cell.alignment = Alignment(horizontal="center")

    out_r = 2
    for mid, grp, urls in parsed_rows:
        for order, url in enumerate(urls, 1):
            if not url:
                continue
            ws_out.cell(out_r, 1, "Create/Edit")
            ws_out.cell(out_r, 2, mid)
            ws_out.cell(out_r, 3, grp)
            ws_out.cell(out_r, 4, url_fn.get(url, fallback_from_url(url)))
            ws_out.cell(out_r, 5, order)
            out_r += 1

    # Auto-fit columns
    for col in ws_out.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws_out.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 60)

    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="ImageStack_Import_Template.xlsx"
    )


# ────────────────────────────────────────────────────────────────────────────
#  ENTRY POINT
# ────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"Pattern Data Processing Backend — http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=False)
